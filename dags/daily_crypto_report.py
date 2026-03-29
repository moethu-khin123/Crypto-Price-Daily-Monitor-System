import sys
sys.path.append("/opt/airflow")

from datetime import datetime, timedelta
import logging
import pandas as pd
import requests

from airflow import DAG
from airflow.operators.python import PythonOperator
from airflow.models import Variable

from scripts.db import get_connection

logger = logging.getLogger(__name__)


# 🚨 FAILURE ALERT
def task_failure_alert(context):
    try:
        token = Variable.get("telegram_bot_token", default_var=None)
        chat_id = Variable.get("telegram_chat_id", default_var=None)

        if not token or not chat_id:
            return

        message = f"""❌ DAG FAILED
DAG: {context['dag'].dag_id}
Task: {context['task_instance'].task_id}
Time: {context['execution_date']}
"""

        requests.post(
            f"https://api.telegram.org/bot{token}/sendMessage",
            data={"chat_id": chat_id, "text": message}
        )

    except Exception as e:
        logger.error(f"Failure alert error: {e}")


# 📊 TASK 1: CALCULATE METRICS 
def calculate_metrics(**context):
    conn = get_connection()
    cur = conn.cursor()

    query = """
        SELECT 
            coin_id,
            DATE(source_last_updated_at) as report_date,
            MIN(price_usd),
            MAX(price_usd),
            AVG(price_usd)
        FROM crypto_prices
        WHERE DATE(source_last_updated_at) = CURRENT_DATE
        GROUP BY coin_id, DATE(source_last_updated_at)
    """

    cur.execute(query)
    rows = cur.fetchall()

    if not rows:
        raise Exception("No data found for today")

    results = []

    for row in rows:
        coin_id, report_date, min_price, max_price, avg_price = row

        results.append({
            "coin_id": coin_id,
            "report_date": str(report_date),
            "min_price": float(min_price),
            "max_price": float(max_price),
            "avg_price": float(avg_price)
        })

    cur.close()
    conn.close()

    logger.info(f"✅ Metrics calculated: {results}")

    context['ti'].xcom_push(key="metrics", value=results)

    return results


# 📁 TASK 2: GENERATE EXCEL
def generate_excel(**context):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.formatting.rule import CellIsRule

    conn = get_connection()

    query = """
        SELECT 
            coin_id,
            DATE(source_last_updated_at) as date,
            AVG(price_usd) as price
        FROM crypto_prices
        WHERE DATE(source_last_updated_at) >= CURRENT_DATE - INTERVAL '7 days'
        GROUP BY coin_id, DATE(source_last_updated_at)
        ORDER BY date
    """

    df = pd.read_sql(query, conn)
    conn.close()

    if df.empty:
        raise Exception("No data for last 7 days")

    # 📊 Split BTC / ETH
    btc = df[df['coin_id'] == 'bitcoin'].copy()
    eth = df[df['coin_id'] == 'ethereum'].copy()

    def process(df):
        df = df.sort_values('date')
        df['Daily Change %'] = df['price'].pct_change() * 100
        df['7-Day Avg'] = df['price'].rolling(7).mean()
        df.rename(columns={
            'date': 'Date',
            'price': 'Price (USD)'
        }, inplace=True)
        return df

    btc = process(btc)
    eth = process(eth)

    # 📊 Summary Sheet
    summary = pd.merge(
        btc[['Date', 'Price (USD)', 'Daily Change %']],
        eth[['Date', 'Price (USD)', 'Daily Change %']],
        on='Date',
        suffixes=(' BTC', ' ETH')
    )

    summary.columns = [
        'Date', 'BTC Price', 'BTC Change %',
        'ETH Price', 'ETH Change %'
    ]

    # 📁 File name
    today = datetime.now().strftime('%Y%m%d')
    file_path = f"/tmp/crypto_report_{today}.xlsx"

    # ✍️ Write Excel
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        btc.to_excel(writer, sheet_name='BTC', index=False)
        eth.to_excel(writer, sheet_name='ETH', index=False)
        summary.to_excel(writer, sheet_name='Summary', index=False)

    # 🎨 Styling
    wb = load_workbook(file_path)

    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # ✅ Header style
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill

        # ✅ Auto column width
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[col_letter].width = max_length + 2

        # ✅ Conditional formatting
        for col in ws.iter_cols(1, ws.max_column):
            if col[0].value and "Change %" in col[0].value:
                col_letter = col[0].column_letter

                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws.max_row}",
                    CellIsRule(operator='greaterThan', formula=['0'], fill=green_fill)
                )

                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws.max_row}",
                    CellIsRule(operator='lessThan', formula=['0'], fill=red_fill)
                )

    wb.save(file_path)

    logger.info(f"📁 Excel generated: {file_path}")

    context['ti'].xcom_push(key="file_path", value=file_path)

    return file_path


# 📩 TASK 3: SEND TELEGRAM
def send_report(**context):
    token = Variable.get("telegram_bot_token", default_var=None)
    chat_id = Variable.get("telegram_chat_id", default_var=None)

    if not token or not chat_id:
        logger.warning("Telegram not configured")
        return

    conn = get_connection()

    query = """
        SELECT 
            coin_id,
            DATE(source_last_updated_at) as date,
            AVG(price_usd) as price
        FROM crypto_prices
        WHERE DATE(source_last_updated_at) >= CURRENT_DATE - INTERVAL '8 days'
        GROUP BY coin_id, DATE(source_last_updated_at)
        ORDER BY coin_id, date
    """

    df = pd.read_sql(query, conn)
    conn.close()

    report_lines = []
    anomalies = []

    def analyze_coin(df_coin):
        df_coin = df_coin.sort_values('date')

        df_coin['change_pct'] = df_coin['price'].pct_change() * 100
        df_coin['ma7'] = df_coin['price'].rolling(7, min_periods=1).mean()

        latest = df_coin.iloc[-1]
        prev = df_coin.iloc[-2]

        latest_price = latest['price']
        daily_change = latest['change_pct']

        if latest['ma7'] > prev['ma7']:
            trend = "⬆ Up"
        elif latest['ma7'] < prev['ma7']:
            trend = "⬇ Down"
        else:
            trend = "➡ Stable"

        return latest_price, daily_change, trend

    for coin_name, coin_label in [("bitcoin", "BTC"), ("ethereum", "ETH")]:
        coin_df = df[df['coin_id'] == coin_name]

        if len(coin_df) < 2:
            continue

        price, change, trend = analyze_coin(coin_df)

        line = f"{coin_label}:  ${price:,.2f}  |  {change:+.2f}%  |  7d: {trend}"
        report_lines.append(line)

        if abs(change) >= 5:
            anomalies.append(f"⚠ {coin_label} abnormal change: {change:+.2f}%")

    today = datetime.now().strftime('%Y-%m-%d')

    message = f"📊 Daily Crypto Report — {today}\n"
    message += "──────────────────────────────\n"

    for line in report_lines:
        message += line + "\n"

    message += "──────────────────────────────\n"

    if anomalies:
        message += "\n".join(anomalies)
    else:
        message += "⚠ No anomalies detected."

    # 📩 Send message
    requests.post(
        f"https://api.telegram.org/bot{token}/sendMessage",
        data={"chat_id": chat_id, "text": message}
    )

    # 📎 Send Excel
    file_path = context['ti'].xcom_pull(
        task_ids='generate_excel',
        key='file_path'
    )

    if file_path:
        with open(file_path, "rb") as f:
            requests.post(
                f"https://api.telegram.org/bot{token}/sendDocument",
                data={"chat_id": chat_id},
                files={"document": f}
            )

    logger.info("✅ Telegram report sent")

# 🧠 DAG
with DAG(
    dag_id="daily_crypto_report",
    start_date=datetime(2024, 1, 1),
    schedule="@daily",
    catchup=False,

    default_args={
        "retries": 2,
        "retry_delay": timedelta(minutes=5),
        "on_failure_callback": task_failure_alert,
    },

    tags=["crypto", "report"],
) as dag:

    t1 = PythonOperator(
        task_id="calculate_metrics",
        python_callable=calculate_metrics,
    )

    t2 = PythonOperator(
        task_id="generate_excel",
        python_callable=generate_excel,
    )

    t3 = PythonOperator(
        task_id="send_report",
        python_callable=send_report,
    )

    t1 >> t2 >> t3